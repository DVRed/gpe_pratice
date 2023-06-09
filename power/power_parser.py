import datetime
import openpyxl
import pandas as pd
from pathlib import Path
import re


class DataRow:
    """
    class to set the data row
    """
    def __init__(self, date: datetime, product: str, hub: str, price: float, bool_product_type: bool):
        self.date = date
        self.product = product
        if hub.find('/') != -1:
            self.hub = self.rename(hub[:hub.find('/')])
            self.hub2 = self.rename(hub[hub.find('/') + 1:])
        else:
            self.hub = self.rename(hub)
            self.hub2 = None
        self.price = price
        self.prices_name = self.rename(hub.replace('/', '_'))
        self.unit = 'MWh'
        self.currency = 'EUR'
        self.id_source = 8
        self.price_type = 'BID' if bool_product_type else 'ASK'
        self.product_types = {'Daily': r'\A(WD|DA|WE|Tue|Wed|Thu|Fri|Sat|Sun|Mon)$',
                              'Week': r'\AWk\d{2}$',
                              'Month': r'\A(BOM|(FEB|MAR|APR|MAY)\d{2})$',
                              'Quarter': r'\AQ\d{3}$',
                              'Season': r'\A(SUM|WIN)\d{2}$',
                              'Year': r'\ACAL\d{2}$'}
        self.product_type = self.get_product_type()

    @staticmethod
    def rename(string_to_rename: str):
        """
        function to rename data for DataFrame
        """
        currency_index = string_to_rename.find('€')
        spread_index = string_to_rename.find('spread')
        if spread_index == -1:
            spread_index = string_to_rename.find('Spread')
        if currency_index != -1:
            return string_to_rename[:currency_index].strip().replace(' ', '_')
        elif spread_index != -1:
            return string_to_rename[:spread_index].replace(' ', '')

    def get_product_type(self):
        """
        function to get product type from product name
        ---
        if ValueError is raised -> there is no mask for such product --> need to add new mask in self.product_types
        """
        price_type = None
        for key in self.product_types.keys():
            if re.search(self.product_types[key], self.product):
                price_type = key
                break
        if price_type is not None:
            return price_type
        else:
            raise ValueError('no mask for product ' + self.product)

    def set_df_row(self):
        """
        function to set the data row as dictionary
        """
        df_row = {
            'date': self.date,
            'prices_name': self.prices_name,
            'price': self.price,
            'hub': self.hub,
            'hub2': self.hub2,
            'unit': self.unit,
            'currency': self.currency,
            'price_type': self.price_type,
            'products': self.product,
            'id_source': self.id_source,
            'product_type': self.product_type
        }
        return df_row


class POWERSheet:
    """
    class to set and read the xlsx sheet
    """
    def __init__(self, sheet_):
        self.date = datetime.datetime.strptime(sheet_.title, '%d%m%Y')
        self.df_sheet = pd.DataFrame(columns=['date', 'prices_name', 'price', 'hub', 'hub2', 'unit', 'currency', 'price_type',
                                              'products', 'id_source', 'product_type'])

        def read_sheet():
            """
            function to read the xlsx sheet data
            and
            put data to DataFrame
            """
            def find_first_col():
                """
                function to find first column with data
                """
                for col_ in sheet_.iter_cols(1, sheet_.max_column):
                    for row_ in range(0, sheet_.max_row):
                        value_ = col_[row_].value
                        if value_ is not None:
                            return col_[row_].column

            first_col = find_first_col()

            def get_stop_rows():
                """
                function to get stop_rows - numbers of rows where to stop reading rows with product names
                sign to stop – colored row
                """
                stop_rows = []
                for row_ in sheet_.iter_rows(1, sheet_.max_row):
                    cell = row_[first_col - 1]
                    if cell.fill.start_color.index == 'FFEF8D4B':
                        stop_rows.append(cell.row)
                return stop_rows

            stop_rows = get_stop_rows()

            def find_first_row(first_row_to_find_not_empty):
                """
                 function to find first row with data from first column with data
                """
                for row_ in sheet_.iter_rows(first_row_to_find_not_empty, sheet_.max_row):
                    value_ = row_[first_col - 1].value
                    if value_ is None:
                        continue
                    elif len(value_.strip()) == 0:
                        continue
                    else:
                        return row_[first_col - 1].row

            first_row_to_find_not_empty = 1

            for i in range(len(stop_rows)):
                stop_row = stop_rows[i]
                first_row = find_first_row(first_row_to_find_not_empty)
                hub_row = first_row - 2

                for row_ in sheet_.iter_rows(first_row, sheet_.max_row):
                    if row_[first_col - 1].row == stop_row:
                        first_row_to_find_not_empty = stop_row + 1
                        break
                    else:
                        product = row_[first_col - 1].value.strip()
                        for col_ in sheet_.iter_cols(2, sheet_.max_column):
                            if col_[hub_row - 1].value is not None:
                                hub = col_[hub_row - 1].value.strip()
                                price_bid = row_[col_[hub_row - 1].column - 1].value
                                price_ask = row_[col_[hub_row - 1].column].value
                                if price_bid is None:
                                    continue
                                else:
                                    new_data_row = DataRow(self.date, product, hub, price_bid, True)
                                    self.df_sheet.loc[len(self.df_sheet.index)] = new_data_row.set_df_row()
                                if price_ask is None:
                                    continue
                                else:
                                    new_data_row = DataRow(self.date, product, hub, price_ask, False)
                                    self.df_sheet.loc[len(self.df_sheet.index)] = new_data_row.set_df_row()

        read_sheet()


class POWERParser:
    """
    class to read the xlsx file
    """
    def __init__(self, file_name):
        self.file_name = file_name
        self.xlsx_file = Path('', self.file_name)
        self.df = pd.DataFrame(columns=['date', 'prices_name', 'price', 'hub', 'hub2', 'unit', 'currency', 'price_type',
                                        'products', 'id_source', 'product_type'])

        self.get_sheets_from_file()

    def get_sheets_from_file(self):
        """
        function to read all sheets from xlsx file
        and
        put data to DataFrame
        """
        wb = openpyxl.load_workbook(self.xlsx_file, data_only=True)
        for sheet_ in wb.worksheets:
            new_GAS_sheet = POWERSheet(sheet_)
            self.df = pd.concat([self.df, new_GAS_sheet.df_sheet], sort=False, axis=0)
            self.df.reset_index(drop=True, inplace=True)

    def concat_two_df(self, df2):
        """
        function to join two DataFrames
        """
        new_df = pd.concat([self.df, df2], sort=False, axis=0)
        new_df.reset_index(drop=True, inplace=True)
        return new_df

    def write_df_to_csv(self):
        """
        function to write DataFrame to .csv
        """
        self.df.to_csv(self.xlsx_file.stem + '.csv')


if __name__ == '__main__':
    file_name = 'ClosingDayPricesPOWER2023.xlsx'
    new_parser = POWERParser(file_name)
    new_parser.write_df_to_csv()



