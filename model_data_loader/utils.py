import re
from io import StringIO
from string import ascii_uppercase as auc
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl.reader.excel import load_workbook


# *******************************
# *         EXCEL UTILS         *
# *******************************

def excel_to_int(excel_index: str) -> int:
    """
    Преобразует индекс колокни Excel (А, B, C, ..., Z, AA, AB, ...) к int индексу.
    Индексу А соответсвует индекс 0.
    """
    num = 0
    for char in excel_index:
        num = num * 26 + (ord(char.upper()) - ord('A')) + 1
    return num - 1


def int_to_excel_index(num: int) -> str:
    """
    Преобразует int индекс к индексу колокни Excel (А, B, C, ..., Z, AA, AB, ...).
    Индексу 0 соответсвует индекс А.
    """
    if num > len(auc) - 1:
        return str(auc[num // 26 - 1]) + auc[num % 26]
    else:
        return auc[num % 26]


def load_xlsx_row(wb, sheet_name, row) -> list[str]:
    return load_xlsx_row_with_skip_list(wb, sheet_name, row, [])


def load_xlsx_row_with_skip_list(wb, sheet_name, row, skip_list) -> list[str]:
    """
    Parameters:
        wb: открытая книга workbook openoyxl
        sheet_name: название листа в книге
        row: номер ряда (нумерация с 1)
        skip_list: список подстрок, если значение в ряду содержит хотя бы одну подстроку из списка, оно будет пропущено
                    (на его месте будет пустая строка)
    Returns:
        Лист строк, содержащий значения ряда, которые не содержат ни одной подстроки из skip_list.
    """
    sheet_ranges = wb[sheet_name]
    if sheet_ranges is None:
        raise ValueError('Target sheet with name "' + sheet_name + '" was not found')
    res = []
    for cell in sheet_ranges[row]:
        line = cell.value
        if any(ext in str(line or '') for ext in skip_list):
            continue
        res.append(str(line or '').strip())
    return res


def get_column_offset(workbook, sheet_name, check_line_index) -> int:
    """
    Возвращает количество первых пустых ячеек в строке.
    Parameters:
        workbook: открытая книга openpyxl
        sheet_name: имя листа
        check_line_index: индекс строки, которая гарантированно не будет полностью пустой
    """
    i = 0
    while workbook[sheet_name][check_line_index][i].value is None:
        i += 1
    return i


def get_first_num_row_index(workbook, sheet_name, check_column_index) -> int:
    """
    Возвращает индекс первой строки с числовым или формульным форматом (нумерация с 0)
    Parameters:
        workbook: открытая книга openpyxl
        sheet_name: имя листа
        check_column_index: индекс колонки, которая гарантированно не будет полностью пустой
    """
    i = 1
    cell = workbook[sheet_name][i][check_column_index]
    while cell.value is None or (cell.data_type != 'n' and cell.data_type != 'f'):
        i += 1
        cell = workbook[sheet_name][i][check_column_index]
    return i - 1


# TODO возможно решение уже есть в openpyxl:
#  https://openpyxl.readthedocs.io/en/stable/api/openpyxl.workbook.external_link.external.html
# TODO do more testing
def get_connections(extracted_zip: dict[str, bytes]) -> dict[str, str]:
    """
    Достает SQL запросы, объявленные в xlsx в разделе "Запросы и подключения"
    Возвращает словарик, в котором
        ключ: имя, которое упоминается в закладке "Область использования" в свойствах подключения
        значение: SQL запрос
    Parameters:
        extracted_zip: словарь, ключи - имена файлов внутри zip'а, значения - бинарные файлы, см. :fun: `extract_zip`
    """
    res = {}

    connections_xml_str = extracted_zip['xl/connections.xml'].decode("utf-8")
    connections_tree = ElementTree.fromstring(connections_xml_str)
    connections_namespace = \
        dict([node for _, node in ElementTree.iterparse(StringIO(connections_xml_str), events=['start-ns'])])
    connections = connections_tree.findall('.//{' + connections_namespace[''] + '}connection')

    table_list = [filename for filename in extracted_zip.keys() if re.match(r'xl/tables/table\d+\.xml', filename)]

    for table_name in table_list:
        table_xml_str = extracted_zip[table_name].decode("utf-8")
        table_xml = ElementTree.fromstring(table_xml_str)
        for connection in connections:
            if connection.attrib['id'] == table_xml.attrib['id']:
                res[table_xml.attrib['name']] = \
                    connection.find('.//{' + connections_namespace[''] + '}dbPr').attrib['command']
    return res


# ****************************
# *         ZIP UTILS        *
# ****************************

def extract_zip(input_zip) -> dict[str, bytes]:
    input_zip = ZipFile(input_zip)
    return {name: input_zip.read(name) for name in input_zip.namelist()}


def zip_back(zip_path: str, files: dict[str, bytes]):
    with open(zip_path, 'w+') as file_created:  # file creation TODO Create file with ZipFile
        pass
    with ZipFile(zip_path, mode='w') as archive:
        for file_name, file_content in files.items():
            archive.writestr(file_name, file_content)


# *****************************
# *         TEXT UTILS        *
# ******************************

def indent_with_tabs(text: str, num: int) -> str:
    """
    Вставляет num табуляций в начало каждой строки
    """
    indent_space = '\t' * num
    text = indent_space + text.replace('\n', '\n' + indent_space)
    return text


if __name__ == '__main__':

    # print(get_column_offset(load_workbook('src/test.xlsx'), 'temp', 10))
    print(get_first_num_row_index(load_workbook('src/BE_SD_copy.xlsx', data_only=False), 'Daily', 10))

    # Testing of 'excel_to_int' function
    print(' BA: ' + str(excel_to_int('BA')))
    print(' CY: ' + str(excel_to_int('CY')))
    print(' AB: ' + str(excel_to_int('AB')))
    print('  A: ' + str(excel_to_int('A')))
    print('  Z: ' + str(excel_to_int('Z')))
    # Testing of 'int_to_excel_index' function
    print(' 52: ' + str(int_to_excel_index(52)))
    print('102: ' + str(int_to_excel_index(102)))
    print(' 27: ' + str(int_to_excel_index(27)))
    print('  0: ' + str(int_to_excel_index(0)))
    print(' 25: ' + str(int_to_excel_index(25)))
