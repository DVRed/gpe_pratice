import re
import shutil
from io import StringIO
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from pandas import DataFrame
import xml.etree.ElementTree as ElementTree
from model_data_loader.db_loader import execute_query_to_dataframe
from model_data_loader.formula_parser import FormulaParser
from model_data_loader.query_generator import generate_query
from model_data_loader.utils import load_xlsx_row, extract_zip, zip_back, get_column_offset, \
    get_first_num_row_index, get_connections
from datetime import datetime

# https://foss.heptapod.net/openpyxl/openpyxl/-/issues/2019
# TODO когда issue закроется, то можно будет сохранять конечный
#  файл с помощью openpyxl, не теряя "Запросы и подключения"


def update_sheet(workbook: Workbook, sheet_name: str, data: DataFrame, column_names: list[str], formulas: list[str],
                 row_offset: int, column_offset: int):
    """
    Заносит данные из data pandas DataFrame'a в лист sheet_name книги workbook,
    создает новую строку в конце, в которую сохраняет формулы из formulas
    Parameters:
        workbook: открытая книга openpyxl, по окончании не закрывается
        sheet_name: лист, в который пишем
        data: pandas DataFrame, названия колонок такие же как и в column_names (кол-во м.б. и меньше)
        column_names: названия всех колонок в листе sheet_name
        formulas: формулы для сохранения в последней строке
        row_offset: начальный адрес строки, нумерация с 0
        column_offset: начальный адрес колонки, нумерация с 0
    """
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)
    for (column_name, column_data) in data.iteritems():
        column_index = column_names.index(column_name) + column_offset
        for j, value in enumerate(column_data):
            if column_name == 'Date':
                # TODO дата печатается как число
                # j+1 т.к. в екселе строки нумеруются с 1
                sheet.cell(row=j + 1 + row_offset, column=column_index).number_format = 'dd.mm.yyyy'
                sheet.cell(row=j + 1 + row_offset, column=column_index).value = value
                # insert_value = value.strftime('%d.%m.%Y')
            else:
                insert_value = value
                sheet.cell(row=j + 1 + row_offset, column=column_index).value = insert_value
    last_row_index = sheet.max_row + 1
    print(last_row_index)
    sheet.insert_rows(last_row_index)
    for column_index, formula in enumerate(formulas):
        sheet.cell(row=last_row_index, column=column_index + column_offset).value = formula


# TODO handle if end_date < begin_date
def main(input_path, temp_path, output_path, sheet_name, begin_date: datetime, end_date: datetime):
    """
    Parameters:
        input_path: путь к входному xlsx файлу
        temp_path: путь для хранения временной копии xlsx файла, включая имя
        output_path: путь к выходному xlsx файлу, включая имя
        sheet_name: имя листа
        begin_date:
        end_date:
    """
    # создаем копию книги
    shutil.copy(input_path, temp_path)
    workbook = load_workbook(temp_path)

    column_offset = get_column_offset(workbook, sheet_name, 10)
    row_offset = get_first_num_row_index(workbook, sheet_name, 10)

    # читаем формулы из последней строки TODO постоянный инкремент последней строки
    formulas = load_xlsx_row(workbook, sheet_name, workbook[sheet_name].max_row)

    # адрес строки с именами колонок row_offset-1 и +1 тк нумерация строк с 1
    column_names = load_xlsx_row(workbook, sheet_name, row_offset)

    # разархивируем копию in-memory
    temporary_extracted = extract_zip(temp_path)

    # парсим формулы екселя в python объекты
    formula_parser = FormulaParser(formulas[:], get_connections(temporary_extracted))

    # генерируем запрос, результат запроса помещаем в pandas DataFrame
    query = generate_query(
        list(formula_parser.data_sources.values()), formula_parser.sum_if_formulas, column_names, begin_date, end_date
    )
    print(query)

    df_generated: DataFrame = execute_query_to_dataframe(query)
    print(df_generated)

    # В копию книги в лист 'sheet_name' вносим изменения
    update_sheet(workbook, sheet_name, df_generated, column_names, formulas, row_offset, column_offset)
    workbook.save(temp_path)

    # разархивируем in-memory только что сохранненый xlsx
    temporary_extracted = extract_zip(temp_path)

    # Копию листа из временной книги вставляем в исходную
    workbook_xml_str = temporary_extracted['xl/workbook.xml'].decode("utf-8")

    # взято с https://stackoverflow.com/a/42338368
    namespace = dict([node for _, node in ElementTree.iterparse(StringIO(workbook_xml_str), events=['start-ns'])])

    tree = ElementTree.fromstring(workbook_xml_str)
    sheets = tree.findall('.//{' + namespace[''] + '}sheet')

    target_xml = None
    target_xml_name = ''
    for sheet in sheets:
        if sheet.attrib['name'] == sheet_name:
            sheet_r_id = sheet.attrib['{' + namespace['r'] + '}id']
            m = re.search(r'\d+$', sheet_r_id)
            if not m:
                raise ValueError('cant find index')
            sheet_index = m.group()
            target_xml_name = 'xl/worksheets/sheet' + sheet_index + '.xml'
            print(target_xml_name)
            target_xml = temporary_extracted[target_xml_name]
    if target_xml is None:
        raise ValueError('Target list "Daily" was not found')

    # разархивируем входной файл, чтобы внести изменения в него
    main_extracted = extract_zip(input_path)
    main_extracted[target_xml_name] = target_xml

    zip_back(output_path, main_extracted)


if __name__ == '__main__':
    t1 = datetime.now()
    main(
        input_path='src/BE SD.xlsx', temp_path='tmp/BE_SD_copy_2.xlsx', output_path='out/res2.xlsx',
        sheet_name='Daily',
        begin_date=datetime(day=1, month=1, year=2015), end_date=datetime(day=1, month=4, year=2024)  # 1 <= month <= 12
    )
    print('executed for ' + str((datetime.now() - t1).seconds))
